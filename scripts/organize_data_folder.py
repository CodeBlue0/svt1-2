#!/usr/bin/env python3
from __future__ import annotations

import csv
import io
import re
import shutil
import subprocess
import tempfile
import unicodedata
import zipfile
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ImportError:  # pragma: no cover - dependency is declared in requirements.txt.
    openpyxl = None

try:
    from numbers_parser import Document
except ImportError:  # pragma: no cover - dependency is declared in requirements.txt.
    Document = None

ROOT = Path(__file__).resolve().parents[1]
EXTRACTED_DIR = ROOT / "extracted"
OUTPUT_DIR = ROOT / "organized_data"
MARKER = OUTPUT_DIR / ".generated_by_organize_data_folder"
DATE_RE = re.compile(r"20\d{2}-\d{2}-\d{2}")
NAME_RE = re.compile(r"^(?P<name>.*?)(?P<short_id>\d{6})_(?P<upload_id>\d+)_(?P<record_id>\d+)_")
SKIP_NAMES = {".DS_Store"}
ARCHIVE_SUFFIXES = set()
DATA_SUFFIXES = {".csv", ".xlsx", ".numbers"}
UNSUPPORTED_SUFFIXES = {".txt", ".crdownload", ""}


def nfc(value: object) -> str:
    return unicodedata.normalize("NFC", str(value or "")).strip()


def safe_name(value: str, fallback: str = "file") -> str:
    value = nfc(value)
    value = re.sub(r"[\\/:*?\"<>|]+", "_", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value or fallback


def should_skip_name(name: str) -> bool:
    path = Path(name)
    return (
        path.name in SKIP_NAMES
        or path.name.startswith("._")
        or "__MACOSX" in path.parts
    )


def date_for(path: Path) -> str:
    rel = nfc(str(path.relative_to(ROOT)))
    match = DATE_RE.search(rel)
    return match.group(0) if match else "undated"


def date_for_text(text: str) -> str:
    match = DATE_RE.search(nfc(text))
    return match.group(0) if match else "undated"


def decode_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr", "utf-16"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            pass
    return data.decode("utf-8", errors="replace")


def spreadsheet_rows_to_csv_bytes(rows: list[list[object]]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def xlsx_bytes_to_csv_bytes(data: bytes) -> bytes:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to convert XLSX files")
    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    return spreadsheet_rows_to_csv_bytes(rows)


def numbers_file_to_csv_bytes(path: Path) -> bytes:
    if Document is None:
        raise RuntimeError("numbers-parser is required to convert Numbers files")
    document = Document(str(path))
    for sheet in document.sheets:
        for table in sheet.tables:
            if table.num_rows == 0:
                continue
            rows = [
                [table.cell(row, col).value for col in range(table.num_cols)]
                for row in range(table.num_rows)
            ]
            return spreadsheet_rows_to_csv_bytes(rows)
    return b""


def data_as_csv_bytes(source: Path, display_name: str, data: bytes | None = None) -> bytes:
    suffix = Path(display_name).suffix.lower()
    if suffix == ".csv":
        if data is not None:
            return data
        return source.read_bytes()
    if suffix == ".xlsx":
        if data is None:
            data = source.read_bytes()
        return xlsx_bytes_to_csv_bytes(data)
    if suffix == ".numbers":
        if data is None and source.suffix.lower() == ".numbers":
            return numbers_file_to_csv_bytes(source)
        with tempfile.NamedTemporaryFile(suffix=".numbers") as tmp:
            tmp.write(data or b"")
            tmp.flush()
            return numbers_file_to_csv_bytes(Path(tmp.name))
    raise ValueError(f"Unsupported data suffix: {suffix}")


def csv_output_name(original_name: str) -> str:
    path = Path(original_name)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return path.name
    if suffix in DATA_SUFFIXES:
        return f"{path.stem}.csv"
    return path.name


def filename_participant_label(name: str) -> str:
    match = NAME_RE.match(nfc(Path(name).name))
    if not match:
        return ""
    person = nfc(match.group("name"))
    short_id = nfc(match.group("short_id"))
    if person and short_id:
        return safe_name(f"{person}_{short_id}", "unknown_participant")
    return safe_name(person or short_id, "unknown_participant")


def filename_identity_parts(name: str) -> tuple[str, str]:
    match = NAME_RE.match(nfc(Path(name).name))
    if not match:
        return "", ""
    return nfc(match.group("name")), nfc(match.group("short_id"))


def clean_participant_id(value: object) -> str:
    cleaned = nfc(value)
    if not cleaned or cleaned.casefold() in {"anonymous", "participant"}:
        return ""
    return safe_name(cleaned, "unknown_participant")


def numeric_student_like_id(value: object) -> str:
    cleaned = nfc(value)
    return cleaned if cleaned.isdigit() and len(cleaned) >= 6 else ""


def usable_text_participant_id(value: object) -> str:
    cleaned = nfc(value)
    if not cleaned or cleaned.casefold() in {"anonymous", "participant"}:
        return ""
    if cleaned.isdigit() and len(cleaned) >= 6:
        return ""
    return cleaned


def usable_text_student_id(value: object) -> str:
    cleaned = nfc(value)
    if not cleaned or cleaned.casefold() in {"anonymous", "participant"}:
        return ""
    if cleaned.isdigit() and len(cleaned) >= 6:
        return ""
    return cleaned


def identity_aliases(student_id: str, participant_id: str, name_from_filename: str, short_id: str, name_unique: bool = False) -> list[str]:
    aliases: list[str] = []
    student_numeric = numeric_student_like_id(student_id)
    participant_numeric = numeric_student_like_id(participant_id)
    participant_text = usable_text_participant_id(participant_id)
    student_text = usable_text_student_id(student_id)
    if student_numeric:
        aliases.append(f"student_numeric:{student_numeric}")
    if participant_numeric:
        aliases.append(f"student_numeric:{participant_numeric}")
    if participant_text:
        aliases.append(f"participant_text:{participant_text.casefold()}")
    if student_text:
        aliases.append(f"student_text:{student_text.casefold()}")
        aliases.append(f"filename_name_unique:{student_text}")
    if name_from_filename and short_id:
        aliases.append(f"filename_name_short:{name_from_filename}:{short_id}")
    if name_from_filename and name_unique:
        aliases.append(f"filename_name_unique:{name_from_filename}")
    return aliases


class UnionFind:
    def __init__(self) -> None:
        self.parent: dict[str, str] = {}

    def find(self, item: str) -> str:
        self.parent.setdefault(item, item)
        if self.parent[item] != item:
            self.parent[item] = self.find(self.parent[item])
        return self.parent[item]

    def union(self, left: str, right: str) -> None:
        root_left = self.find(left)
        root_right = self.find(right)
        if root_left == root_right:
            return
        self.parent[max(root_left, root_right)] = min(root_left, root_right)


def canonical_participant_label(aliases: set[str]) -> str:
    participant_texts = sorted(alias.split(":", 1)[1] for alias in aliases if alias.startswith("participant_text:"))
    if participant_texts:
        return safe_name(participant_texts[0], "unknown_participant")
    name_short = sorted(alias.split(":", 1)[1] for alias in aliases if alias.startswith("filename_name_short:"))
    if name_short:
        return safe_name(name_short[0].replace(":", "_"), "unknown_participant")
    student_ids = sorted(alias.split(":", 1)[1] for alias in aliases if alias.startswith("student_numeric:"))
    if student_ids:
        return safe_name(student_ids[0], "unknown_participant")
    student_texts = sorted(alias.split(":", 1)[1] for alias in aliases if alias.startswith("student_text:"))
    if student_texts:
        return safe_name(student_texts[0], "unknown_participant")
    name_unique = sorted(alias.split(":", 1)[1] for alias in aliases if alias.startswith("filename_name_unique:"))
    if name_unique:
        return safe_name(name_unique[0], "unknown_participant")
    return "unknown_participant"


def csv_metadata(data: bytes) -> dict[str, str]:
    try:
        reader = csv.DictReader(io.StringIO(decode_bytes(data)))
        row = next(reader, None)
    except Exception:
        return {}
    if not row:
        return {}
    normalized = {nfc(key).casefold(): nfc(value) for key, value in row.items() if key is not None}
    student_id = nfc(normalized.get("student_id", ""))
    participant_id = nfc(normalized.get("participant_id", ""))
    task_raw = nfc(normalized.get("task", "")).casefold()
    date_raw = " ".join(
        nfc(normalized.get(key, ""))
        for key in ("date", "timestamp", "session_id")
        if normalized.get(key)
    )

    task = ""
    if "maze" in task_raw or "rsvp" in task_raw:
        task = "maze"
    elif "cst" in task_raw or "svt" in task_raw:
        task = "svt"

    return {
        "participant": clean_participant_id(student_id) or clean_participant_id(participant_id),
        "student_id": student_id,
        "participant_id": participant_id,
        "task": task,
        "date": date_for_text(date_raw) if date_raw else "",
    }


def file_metadata(source: Path, display_name: str, data: bytes | None = None, participant_alias_map: dict[str, str] | None = None) -> dict[str, str]:
    text = f"{source.relative_to(EXTRACTED_DIR)}::{display_name}"
    suffix = Path(display_name).suffix.lower()
    path_task = classify_text(text, suffix)
    name_from_filename, short_id = filename_identity_parts(display_name)
    if not name_from_filename and not short_id:
        name_from_filename, short_id = filename_identity_parts(source.name)
    metadata = {
        "participant": filename_participant_label(display_name) or filename_participant_label(source.name) or "unknown_participant",
        "student_id": "",
        "participant_id": "",
        "name_from_filename": name_from_filename,
        "short_id": short_id,
        "task": path_task,
        "date": date_for_text(text),
    }
    if suffix == ".csv":
        if data is None and source.is_file():
            try:
                data = source.read_bytes()
            except OSError:
                data = None
        if data:
            parsed = csv_metadata(data)
            metadata["participant"] = parsed.get("participant") or metadata["participant"]
            metadata["student_id"] = parsed.get("student_id", "")
            metadata["participant_id"] = parsed.get("participant_id", "")
            metadata["task"] = parsed.get("task") or metadata["task"]
            metadata["date"] = parsed.get("date") or metadata["date"]
    aliases = identity_aliases(
        metadata.get("student_id", ""),
        metadata.get("participant_id", ""),
        metadata.get("name_from_filename", ""),
        metadata.get("short_id", ""),
    )
    if participant_alias_map:
        for alias in aliases:
            resolved = participant_alias_map.get(alias)
            if resolved:
                metadata["participant"] = resolved
                break
    if path_task == "excluded_svt_s1":
        metadata["task"] = "excluded_svt_s1"
    if metadata["task"] == "unknown" and suffix not in DATA_SUFFIXES:
        metadata["task"] = "unsupported"
    return metadata


def build_participant_alias_map(files: Iterable[Path]) -> dict[str, str]:
    raw_records: list[dict[str, str]] = []

    def add_record(source: Path, display_name: str, data: bytes | None) -> None:
        metadata = file_metadata(source, display_name, data, None)
        if metadata["task"] == "excluded_svt_s1":
            return
        raw_records.append(metadata)

    for source in files:
        if source.suffix.lower() == ".zip":
            try:
                with zipfile.ZipFile(source) as archive:
                    for info in archive.infolist():
                        if info.is_dir() or should_skip_name(info.filename):
                            continue
                        if Path(info.filename).suffix.lower() == ".csv":
                            add_record(source, info.filename, archive.read(info))
            except zipfile.BadZipFile:
                pass
            continue
        if source.suffix.lower() == ".rar":
            try:
                with tempfile.TemporaryDirectory(prefix="organize_rar_alias_") as tmp:
                    tmp_path = Path(tmp)
                    subprocess.run(
                        ["bsdtar", "-xf", str(source), "-C", str(tmp_path)],
                        check=True,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                    )
                    for extracted in sorted(p for p in tmp_path.rglob("*") if p.is_file() and not should_skip_name(str(p.relative_to(tmp_path)))):
                        if extracted.suffix.lower() == ".csv":
                            add_record(source, str(extracted.relative_to(tmp_path)), extracted.read_bytes())
            except (subprocess.CalledProcessError, FileNotFoundError):
                pass
            continue
        if source.suffix.lower() == ".csv":
            add_record(source, source.name, None)

    name_to_shorts: dict[str, set[str]] = {}
    for record in raw_records:
        name = record.get("name_from_filename", "")
        short = record.get("short_id", "")
        if name and short:
            name_to_shorts.setdefault(name, set()).add(short)

    uf = UnionFind()
    record_aliases: list[list[str]] = []
    for record in raw_records:
        name = record.get("name_from_filename", "")
        short = record.get("short_id", "")
        aliases = identity_aliases(
            record.get("student_id", ""),
            record.get("participant_id", ""),
            name,
            short,
            bool(name) and len(name_to_shorts.get(name, set())) == 1,
        )
        if not aliases:
            continue
        record_aliases.append(aliases)
        first = aliases[0]
        for alias in aliases[1:]:
            uf.union(first, alias)

    root_aliases: dict[str, set[str]] = {}
    for aliases in record_aliases:
        for alias in aliases:
            root_aliases.setdefault(uf.find(alias), set()).add(alias)

    alias_map: dict[str, str] = {}
    for aliases in root_aliases.values():
        label = canonical_participant_label(aliases)
        for alias in aliases:
            alias_map[alias] = label
    return alias_map


def classify_text(text: str, suffix: str = "") -> str:
    lower = nfc(text).casefold()
    if lower.startswith("svt_s1") or "/svt_s1/" in lower:
        return "excluded_svt_s1"
    if "maze" in lower or "rsvp" in lower:
        return "maze"
    if "svt" in lower or "cst" in lower:
        return "svt"
    if suffix.lower() in {".zip", ".rar"}:
        return "unknown"
    return "unknown"


def classify_task(path: Path) -> str:
    rel_lower = nfc(str(path.relative_to(EXTRACTED_DIR))).casefold()
    if rel_lower.startswith("svt_s1"):
        return "excluded_svt_s1"
    text_class = classify_text(rel_lower, path.suffix)
    if text_class != "unknown":
        return text_class
    if path.suffix.lower() == ".csv":
        try:
            header = path.read_text(encoding="utf-8-sig", errors="ignore").splitlines()[0].casefold()
            if "task" in header and "correct" in header and "rt" in header:
                for line in path.read_text(encoding="utf-8-sig", errors="ignore").splitlines()[1:12]:
                    low = line.casefold()
                    if ",cst," in low:
                        return "svt"
                    if "maze" in low or "rsvp" in low:
                        return "maze"
        except Exception:
            pass
    return "unknown"


def bucket_for(path: Path) -> tuple[str, str]:
    suffix = path.suffix.lower()
    task = classify_task(path)
    if suffix in DATA_SUFFIXES:
        return task, date_for(path)
    return "unsupported", suffix.lstrip(".") or "no_extension"


def participant_target_dir(metadata: dict[str, str]) -> Path:
    participant = safe_name(metadata.get("participant") or "unknown_participant", "unknown_participant")
    task = safe_name(metadata.get("task") or "unknown", "unknown")
    date = safe_name(metadata.get("date") or "undated", "undated")
    return OUTPUT_DIR / "participants" / participant / task / date


def bucket_for_zip_member(archive: Path, member_name: str) -> tuple[str, str, str]:
    suffix = Path(member_name).suffix.lower()
    task = classify_text(f"{archive.relative_to(EXTRACTED_DIR)}::{member_name}", suffix)
    if suffix in DATA_SUFFIXES:
        return task, date_for_text(f"{archive.name}/{member_name}"), task
    return "unsupported", suffix.lstrip(".") or "no_extension", task


def bucket_for_archive_member(archive: Path, member_name: str) -> tuple[str, str, str]:
    return bucket_for_zip_member(archive, member_name)


def unique_link_path(directory: Path, original_name: str, used: set[Path]) -> Path:
    base = safe_name(original_name)
    candidate = directory / base
    if candidate not in used and not candidate.exists() and not candidate.is_symlink():
        used.add(candidate)
        return candidate
    stem = Path(base).stem
    suffix = Path(base).suffix
    counter = 2
    while True:
        candidate = directory / f"{stem}__{counter}{suffix}"
        if candidate not in used and not candidate.exists() and not candidate.is_symlink():
            used.add(candidate)
            return candidate
        counter += 1


def unique_file_path(directory: Path, original_name: str, used: set[Path]) -> Path:
    return unique_link_path(directory, original_name, used)


def rebuild_output() -> None:
    if OUTPUT_DIR.exists():
        if not MARKER.exists():
            raise SystemExit(f"Refusing to overwrite non-generated folder: {OUTPUT_DIR}")
        shutil.rmtree(OUTPUT_DIR)
    OUTPUT_DIR.mkdir(parents=True)
    MARKER.write_text("Generated by scripts/organize_data_folder.py; safe to delete/rebuild.\n", encoding="utf-8")


def main() -> None:
    rebuild_output()
    manifest_rows: list[dict[str, str]] = []
    used: set[Path] = set()
    seen_content: set[str] = set()
    files = sorted(path for path in EXTRACTED_DIR.rglob("*") if path.is_file() and not should_skip_name(path.name))
    participant_alias_map = build_participant_alias_map(files)
    for source in files:
        source_metadata = file_metadata(source, source.name, participant_alias_map=participant_alias_map)
        if source_metadata["task"] == "excluded_svt_s1":
            continue
        if source.suffix.lower() == ".zip":
            try:
                with zipfile.ZipFile(source) as archive:
                    for info in archive.infolist():
                        if info.is_dir() or should_skip_name(info.filename):
                            continue
                        if Path(info.filename).suffix.lower() not in DATA_SUFFIXES:
                            continue
                        member_data = archive.read(info)
                        metadata = file_metadata(source, info.filename, member_data, participant_alias_map)
                        if metadata["task"] == "excluded_svt_s1":
                            continue
                        csv_data = data_as_csv_bytes(source, info.filename, member_data)
                        content_hash = __import__("hashlib").sha256(csv_data).hexdigest()
                        if content_hash in seen_content:
                            continue
                        seen_content.add(content_hash)
                        target_dir = participant_target_dir(metadata)
                        target_dir.mkdir(parents=True, exist_ok=True)
                        target_path = unique_file_path(target_dir, csv_output_name(info.filename), used)
                        target_path.write_bytes(csv_data)
                        manifest_rows.append({
                            "participant": metadata["participant"],
                            "task": metadata["task"],
                            "date": metadata["date"],
                            "suffix": ".csv",
                            "organized_path": str(target_path.relative_to(ROOT)),
                            "source_path": f"{source.relative_to(ROOT)}::{info.filename}",
                        })
            except zipfile.BadZipFile:
                target_dir = OUTPUT_DIR / "unsupported" / "bad_zip"
                target_dir.mkdir(parents=True, exist_ok=True)
                link_path = unique_link_path(target_dir, source.name, used)
                rel_target = Path(__import__("os").path.relpath(source, link_path.parent))
                link_path.symlink_to(rel_target)
                manifest_rows.append({
                    "participant": "unknown_participant",
                    "task": "unsupported",
                    "date": date_for(source),
                    "suffix": source.suffix.lower() or "<none>",
                    "organized_path": str(link_path.relative_to(ROOT)),
                    "source_path": str(source.relative_to(ROOT)),
                })
            continue
        if source.suffix.lower() == ".rar":
            try:
                with tempfile.TemporaryDirectory(prefix="organize_rar_") as tmp:
                    tmp_path = Path(tmp)
                    subprocess.run(
                        ["bsdtar", "-xf", str(source), "-C", str(tmp_path)],
                        check=True,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                    )
                    for extracted in sorted(p for p in tmp_path.rglob("*") if p.is_file() and not should_skip_name(str(p.relative_to(tmp_path)))):
                        member_name = str(extracted.relative_to(tmp_path))
                        if extracted.suffix.lower() not in DATA_SUFFIXES:
                            continue
                        member_data = extracted.read_bytes()
                        metadata = file_metadata(source, member_name, member_data, participant_alias_map)
                        if metadata["task"] == "excluded_svt_s1":
                            continue
                        csv_data = data_as_csv_bytes(extracted, member_name, member_data)
                        content_hash = __import__("hashlib").sha256(csv_data).hexdigest()
                        if content_hash in seen_content:
                            continue
                        seen_content.add(content_hash)
                        target_dir = participant_target_dir(metadata)
                        target_dir.mkdir(parents=True, exist_ok=True)
                        target_path = unique_file_path(target_dir, csv_output_name(extracted.name), used)
                        target_path.write_bytes(csv_data)
                        manifest_rows.append({
                            "participant": metadata["participant"],
                            "task": metadata["task"],
                            "date": metadata["date"],
                            "suffix": ".csv",
                            "organized_path": str(target_path.relative_to(ROOT)),
                            "source_path": f"{source.relative_to(ROOT)}::{member_name}",
                        })
            except (subprocess.CalledProcessError, FileNotFoundError) as exc:
                manifest_rows.append({
                    "participant": "unknown_participant",
                    "task": "unsupported",
                    "date": date_for(source),
                    "suffix": source.suffix.lower() or "<none>",
                    "organized_path": "",
                    "source_path": f"{source.relative_to(ROOT)}::EXTRACT_FAILED:{exc.__class__.__name__}",
                })
            continue
        metadata = source_metadata
        if source.suffix.lower() not in DATA_SUFFIXES:
            continue
        target_dir = participant_target_dir(metadata)
        target_dir.mkdir(parents=True, exist_ok=True)
        try:
            csv_data = data_as_csv_bytes(source, source.name)
            content_hash = __import__("hashlib").sha256(csv_data).hexdigest()
        except OSError:
            content_hash = str(source)
        if content_hash in seen_content:
            continue
        seen_content.add(content_hash)
        link_path = unique_link_path(target_dir, csv_output_name(source.name), used)
        if source.suffix.lower() == ".csv":
            rel_target = Path("../" * len(link_path.relative_to(OUTPUT_DIR).parents)) / source.relative_to(ROOT)
            # More reliable relative target from link directory.
            rel_target = Path(__import__("os").path.relpath(source, link_path.parent))
            link_path.symlink_to(rel_target)
        else:
            link_path.write_bytes(csv_data)
        manifest_rows.append({
            "participant": metadata["participant"],
            "task": metadata["task"],
            "date": metadata["date"],
            "suffix": ".csv",
            "organized_path": str(link_path.relative_to(ROOT)),
            "source_path": str(source.relative_to(ROOT)),
        })

    manifest_path = OUTPUT_DIR / "manifest.csv"
    with manifest_path.open("w", newline="", encoding="utf-8-sig") as f:
        fieldnames = ["participant", "task", "date", "suffix", "organized_path", "source_path"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(manifest_rows)

    readme = OUTPUT_DIR / "README.md"
    readme.write_text(
        "# Organized experiment data\n\n"
        "This folder is a generated symlink and unpacked-archive view of `extracted/`.\n"
        "Original raw files remain in `extracted/`; do not edit files through this folder.\n\n"
        "## Layout\n"
        "- `participants/{participant_id}/svt/YYYY-MM-DD/`: SVT/CST files by participant and date.\n"
        "- `participants/{participant_id}/maze/YYYY-MM-DD/`: MAZE/RSVP files by participant and date.\n"
        "- `manifest.csv`: mapping from participant/task/date folders to original source paths.\n\n"
        "Zip and rar archives are unpacked here and the archive files themselves are not kept in this organized view.\n"
        "Spreadsheet inputs (`.xlsx` and `.numbers`) are converted to `.csv` in this organized view.\n"
        "Raw original archives remain under `extracted/`.\n\n"
        "Rebuild with: `python scripts/organize_data_folder.py`\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT_DIR.relative_to(ROOT)} with {len(manifest_rows)} entries")


if __name__ == "__main__":
    main()
